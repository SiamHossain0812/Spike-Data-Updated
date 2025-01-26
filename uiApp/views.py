import csv
import datetime
import openpyxl
import numpy as np
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render
from .models import SpikeData, StationName
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.db import connection
from django.http import JsonResponse

def detect_gaps(data, interval_minutes=15, gap_threshold_points=192):
    """
    Detects gaps in the dataset where 192 or more consecutive data points are missing.
    :param data: List of dictionaries with 'dateTime' and 'value'.
    :param interval_minutes: Expected interval between data points (default 15 minutes).
    :param gap_threshold_points: Number of missing consecutive points to consider as a gap (default 192).
    :return: Tuple (segments, total_data_gaps), where:
             segments: List of segments, each a list of data points.
             total_data_gaps: Total number of missing data points between segments.
    """
    # Sort data by datetime
    data.sort(key=lambda x: x['dateTime'])

    # Generate the full range of expected timestamps
    start_time = data[0]['dateTime']
    end_time = data[-1]['dateTime']
    expected_timestamps = [start_time + timedelta(minutes=i * interval_minutes)
                           for i in range(int((end_time - start_time).total_seconds() / (interval_minutes * 60)) + 1)]
    
    # Identify missing timestamps
    actual_timestamps = {point['dateTime'] for point in data}
    missing_timestamps = [ts for ts in expected_timestamps if ts not in actual_timestamps]
    
    # Detect gaps of consecutive missing timestamps
    gaps = []
    current_gap = []

    for ts in missing_timestamps:
        if not current_gap or ts - current_gap[-1] == timedelta(minutes=interval_minutes):
            current_gap.append(ts)
        else:
            if len(current_gap) >= gap_threshold_points:  # Check if the gap meets the threshold
                gaps.append(current_gap)
            current_gap = [ts]

    # Check the last gap
    if len(current_gap) >= gap_threshold_points:
        gaps.append(current_gap)

    # Split data into segments based on gaps
    segments = []
    previous_end_time = start_time

    for gap in gaps:
        # Collect data points before the gap
        segment = [point for point in data if previous_end_time <= point['dateTime'] < gap[0]]
        if segment:
            segments.append(segment)
        previous_end_time = gap[-1] + timedelta(minutes=interval_minutes)

    # Add the remaining data after the last gap
    last_segment = [point for point in data if point['dateTime'] >= previous_end_time]
    if last_segment:
        segments.append(last_segment)

    # Calculate total data gaps
    total_data_gaps = sum(len(gap) for gap in gaps)

    return segments, total_data_gaps


def process_segments_with_multiple_gaps(data, selected_station_info=None):
    """
    Processes a dataset with multiple gaps by splitting it into segments,
    calculating thresholds, and replacing invalid/abnormal values for each segment.
    :param data: List of data points (e.g., [{'dateTime': ..., 'value': ...}]).
    :param selected_station_info: Station-specific information for threshold adjustments.
    :return: Tuple (processed_data, segment_summaries, abnormal_details).
    """
    # Step 1: Detect gaps and split into segments
    segments, total_data_gaps = detect_gaps(data)

    # Log the total number of segments and total data gaps
    print(f"Total Segments Created: {len(segments)}")
    print(f"Total Data Gaps Between Segments: {total_data_gaps}")

    # Step 2: Process each segment independently
    processed_data = []
    segment_summaries = []
    abnormal_details = []  # Collect abnormal value details across all segments

    for idx, segment in enumerate(segments):
        print(f"\nProcessing Segment {idx + 1}...")

        # Extract values for threshold calculation
        values = [point['value'] for point in segment]

        # Calculate thresholds for this segment
        lower_threshold, upper_threshold = calculate_dynamic_thresholds(values)
        print(f"Segment {idx + 1} Thresholds: Lower={lower_threshold}, Upper={upper_threshold}")

        replaced_values, invalid_count, abnormal_count, segment_abnormal_details, modified_abnormal_count = replace_invalid_values(
            values, selected_station_info
        )

        # Log modified abnormal count for the current segment
        print(f"Segment {idx + 1}: Abnormal values modified: {modified_abnormal_count}")

        # Update processed data with modified values
        for i, point in enumerate(segment):
            processed_data.append({
                'dateTime': point['dateTime'],
                'original_value': point['value'],
                'modified_value': replaced_values[i],
            })

        # Add segment-specific abnormal details
        for detail in segment_abnormal_details:
            # Map index to the original datetime
            detail['dateTime'] = segment[detail['index']]['dateTime']
            abnormal_details.append(detail)

        # Collect segment-specific summary
        segment_summaries.append({
            'segment_number': idx + 1,
            'total_data_points': len(segment),
            'invalid_count': invalid_count,
            'abnormal_count': abnormal_count,
            'lower_threshold': lower_threshold,
            'upper_threshold': upper_threshold,
        })

    # Log segment-specific summaries
    for summary in segment_summaries:
        print(f"\nSummary for Segment {summary['segment_number']}:")
        print(f"  Total Data Points: {summary['total_data_points']}")
        print(f"  Invalid Count: {summary['invalid_count']}")
        print(f"  Abnormal Count: {summary['abnormal_count']}")
        print(f"  Thresholds -> Lower: {summary['lower_threshold']}, Upper: {summary['upper_threshold']}")

    return processed_data, segment_summaries, abnormal_details




def is_invalid(val):
    """Helper function to check if a value is invalid (9999, -9999999, empty or None)."""
    return str(val).startswith('9999') or str(val).startswith('-9999999') or val == '' or val is None

def calculate_dynamic_thresholds(values, lower_percentile=1, upper_percentile=99):
    """
    Calculate dynamic thresholds using percentiles and Median Absolute Deviation (MAD).

    :param values: List or numpy array of values.
    :param lower_percentile: Lower percentile for threshold.
    :param upper_percentile: Upper percentile for threshold.
    :return: Tuple of (lower_threshold, upper_threshold).
    """
    # Remove invalid values
    valid_values = [v for v in values if not is_invalid(v)]

    # Calculate percentiles
    lower_threshold = np.percentile(valid_values, lower_percentile)
    upper_threshold = np.percentile(valid_values, upper_percentile)

    # Calculate Median Absolute Deviation (MAD)
    median = np.median(valid_values)
    mad = 1.4826 * np.median([abs(v - median) for v in valid_values])

    # Adjust thresholds with MAD for robustness
    lower_threshold = max(lower_threshold, median - 2 * mad)
    upper_threshold = min(upper_threshold, median + 2 * mad)

    return lower_threshold, upper_threshold

def is_abnormal_dynamic(values, index, lower_threshold, upper_threshold):
    """
    Check if a value is abnormal based on dynamic thresholds.

    :param values: List or numpy array of values.
    :param index: Index of the value to check.
    :param lower_threshold: Lower dynamic threshold.
    :param upper_threshold: Upper dynamic threshold.
    :return: True if the value is abnormal, False otherwise.
    """
    if is_invalid(values[index]):
        return False

    value = values[index]
    return value < lower_threshold or value > upper_threshold

import numpy as np

def replace_invalid_values(values, selected_station_info=None):
    """
    Replace invalid values (e.g., starts with 9999, -9999999) with 0,
    and replace abnormal values only if the difference from the previous value is greater than 1.
    Abnormal values are replaced with the average of the previous 4 and next 4 valid values.
    Additionally, tracks and returns abnormal value details.

    :param values: List or numpy array of values.
    :param selected_station_info: Station-specific information for threshold adjustments.
    :return: Tuple of (modified_values, invalid_count, abnormal_count, abnormal_details, modified_abnormal_count).
    """
    # Ensure it's a numpy array for easy manipulation
    values = np.array(values, dtype=float)
    
    # Calculate initial dynamic thresholds
    lower_threshold, upper_threshold = calculate_dynamic_thresholds(values)
    print(f"DEBUG: Initial Calculated Thresholds -> Lower: {lower_threshold}, Upper: {upper_threshold}")
    
    # Adjust thresholds based on station info (if provided)
    if selected_station_info:
        try:
            # Extract the station code from the selected_station_info
            station_code = selected_station_info.split("(")[1].split(")")[0]
            print(f"DEBUG: Extracted Station Code: {station_code}")

            # Query for station-specific thresholds in the database
            station_record = StationRecord.objects.filter(station_id=station_code).first()
            if station_record:
                print(f"DEBUG: StationRecord found for {station_code}")
                print(f"DEBUG: Recorded Thresholds -> Highest WL: {station_record.recorded_highest_wl}, "
                      f"Lowest WL: {station_record.recorded_lowest_wl}")

                # Adjust thresholds based on recorded values
                lower_threshold = max(lower_threshold, station_record.recorded_lowest_wl)
                upper_threshold = min(upper_threshold, station_record.recorded_highest_wl)

                print(f"DEBUG: Adjusted Thresholds -> Lower: {lower_threshold}, Upper: {upper_threshold}")
            else:
                print(f"DEBUG: No StationRecord found for {station_code}, using calculated thresholds.")
        except (IndexError, AttributeError):
            print("DEBUG: Invalid station info format or missing fields, skipping station-specific adjustments.")

    print(f"DEBUG: Final Thresholds Used -> Lower: {lower_threshold}, Upper: {upper_threshold}")

    # Create a copy for modifications
    modified_values = values.copy()
    invalid_count = 0  # Track invalid values count
    abnormal_count = 0  # Track abnormal values count
    modified_abnormal_count = 0  # Count how many abnormal values were actually modified
    abnormal_details = []  # Collect details of abnormal values

    # Step 1: Replace invalid values (e.g., starts with 9999/-9999999) with 0
    for i, val in enumerate(modified_values):
        if is_invalid(val):
            modified_values[i] = 0.0  # Replace invalid value with 0
            invalid_count += 1

    # Step 2: Replace abnormal values only if the difference from the previous value > 1
    for i, val in enumerate(modified_values):
        if i == 0:  # Skip the first value since there's no previous value to compare with
            continue

        if is_abnormal_dynamic(modified_values, i, lower_threshold, upper_threshold):
            abnormal_count += 1

            # Compare with the previous value
            if abs(modified_values[i] - modified_values[i - 1]) > 1:
                # Collect previous 4 and next 4 valid values
                prev_values = [v for v in modified_values[max(0, i - 4):i] if not is_invalid(v)]
                next_values = [v for v in modified_values[i + 1:i + 5] if not is_invalid(v)]
                surrounding_values = prev_values + next_values

                # Replace with mean of surrounding valid values
                if surrounding_values:
                    mean = np.mean(surrounding_values)
                    modified_value = float(f"{mean:.3f}")
                    # Add to abnormal details
                    abnormal_details.append({
                        "index": i,
                        "original_value": modified_values[i],
                        "modified_value": modified_value
                    })
                    modified_values[i] = modified_value
                    modified_abnormal_count += 1

    # Ensure all values have 3 digits after decimal before returning
    modified_values = [float(f"{val:.3f}") for val in modified_values]

    # Print abnormal value stats
    print(f"Total abnormal values detected: {abnormal_count}")
    print(f"Total abnormal values modified: {modified_abnormal_count}")

    return modified_values, invalid_count, abnormal_count, abnormal_details, modified_abnormal_count



# Note: This function assumes the existence of helper functions `is_invalid`, 
# `calculate_dynamic_thresholds`, and `is_abnormal_dynamic`.
# These helper functions should be defined elsewhere in the codebase.

def spikedata(request):
    # Retrieve stored values from the session
    last_uploaded_file_name = request.session.get('uploaded_file_name', '')
    stored_start_date = request.session.get('start_date', '')
    stored_end_date = request.session.get('end_date', '')
    stored_rate_of_change = request.session.get('rate_of_change', '')
    stored_station_id = request.session.get('station_id', '')

    # Initialize formatted start and end date variables
    formatted_start_date = ''
    formatted_end_date = ''
    station_name = ''
    selected_station_info = ''

    # Check if the form is submitted with a POST request
    if request.method == 'POST':
        # Get start and end date from the form
        start_date = request.POST.get('start_date', '')
        end_date = request.POST.get('end_date', '')
        rate_of_change = request.POST.get('rate_of_change', '')
        station_id = request.POST.get('station_name', '')

        # Convert start and end date to DD/MM/YYYY format
        try:
            if start_date:
                formatted_start_date = datetime.strptime(start_date, '%Y-%m-%d').strftime('%d/%m/%Y')
            if end_date:
                formatted_end_date = datetime.strptime(end_date, '%Y-%m-%d').strftime('%d/%m/%Y')

            # Store the formatted dates and other values in session
            request.session['start_date'] = formatted_start_date
            request.session['end_date'] = formatted_end_date
            request.session['rate_of_change'] = rate_of_change
            request.session['station_id'] = station_id

        except ValueError:
            return render(request, 'spikedata.html', {'error': 'Invalid date format. Please use YYYY-MM-DD format.'})

        # Retrieve station name by ID from the database
        if station_id:
            station = StationName.objects.filter(id=station_id).first()
            if station:
                station_name = station.station_name
                selected_station_info = f"Station Name: {station_name} ({station_id})"
                print(f"DEBUG: Selected Station Info: {selected_station_info}")

        # Handle file upload if a file is provided
        if 'file_upload' in request.FILES:
            file = request.FILES['file_upload']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            filepath = fs.path(filename)

            # Store the uploaded file name in session
            request.session['uploaded_file_name'] = filename
            all_data = []

            # Handle CSV file processing
            if filename.endswith('.csv'):
                try:
                    with open(filepath, 'r', encoding='utf-8') as csvfile:
                        reader = csv.DictReader(csvfile)
                        for row in reader:
                            dateTime_str = row.get('dateTime', '').strip()
                            value = row.get('value', '').strip()

                            try:
                                value = float(value) if value not in ('-', '', ' ') else None
                            except ValueError:
                                value = None

                            if dateTime_str and value is not None:
                                try:
                                    dateTime_obj = datetime.strptime(dateTime_str, '%d/%m/%Y %H:%M')
                                except ValueError:
                                    try:
                                        dateTime_obj = datetime.strptime(dateTime_str, '%Y-%m-%d %H:%M:%S')
                                    except ValueError:
                                        dateTime_obj = None

                                if dateTime_obj:
                                    all_data.append({'dateTime': dateTime_obj, 'value': value})

                except UnicodeDecodeError:
                    return render(request, 'spikedata.html', {'error': 'Error decoding CSV file. Please check the file encoding and try again.'})

            # Handle Excel file processing
            elif filename.endswith('.xlsx'):
                try:
                    workbook = openpyxl.load_workbook(filepath)
                    for sheet in workbook.worksheets:
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            dateTime_val = row[0] if row[0] is not None else ''
                            value = row[1] if row[1] is not None else ''

                            try:
                                value = float(value) if value not in ('-', '', ' ') else None
                            except ValueError:
                                value = None

                            if dateTime_val and value is not None:
                                try:
                                    dateTime_obj = datetime.strptime(str(dateTime_val), '%d/%m/%Y %H:%M')
                                except ValueError:
                                    try:
                                        dateTime_obj = datetime.strptime(str(dateTime_val), '%Y-%m-%d %H:%M:%S')
                                    except ValueError:
                                        dateTime_obj = None

                                if dateTime_obj:
                                    all_data.append({'dateTime': dateTime_obj, 'value': value})

                except Exception as e:
                    return render(request, 'spikedata.html', {'error': f'Error processing Excel file: {str(e)}'})

            else:
                return render(request, 'spikedata.html', {'error': 'Unsupported file format.'})

            # Detect gaps and process segments
            processed_data, segment_summaries, abnormal_details = process_segments_with_multiple_gaps(all_data, selected_station_info)

            # Log segment details
            print(f"Total Segments Created: {len(segment_summaries)}")
            for summary in segment_summaries:
                print(f"Segment {summary['segment_number']}: Total Points={summary['total_data_points']}, "
                      f"Invalid Count={summary['invalid_count']} "
                      f"Thresholds=[{summary['lower_threshold']}, {summary['upper_threshold']}]")

            # Store processed data back into the SpikeData model
            SpikeData.objects.all().delete()
            for record in processed_data:
                SpikeData.objects.create(
                    dateTime=record['dateTime'],
                    value=record['original_value'],
                    modified_value=record['modified_value']
                )

            # Prepare the summary with the stored values included
            summary = {
                'total_data_points': len(all_data),
                'total_segments': len(segment_summaries),
                'invalid_data_points': sum([summary['invalid_count'] for summary in segment_summaries]),
                'abnormal_data_points': len(abnormal_details),  # Pass total abnormal points
                'last_uploaded_file_name': filename,
                'stored_start_date': formatted_start_date,
                'stored_end_date': formatted_end_date,
                'stored_rate_of_change': rate_of_change,
                'stored_station_name': station_name
            }

            return render(request, 'spikedata.html', {
                'success': True,
                'message': 'File uploaded and data analyzed successfully.',
                'summary': summary,
                'abnormal_details': abnormal_details,  # Pass abnormal details to frontend
                'stations': StationName.objects.all(),
            })

    stations = StationName.objects.all()
    return render(request, 'spikedata.html', {
        'last_uploaded_file_name': last_uploaded_file_name,
        'summary': {
            'last_uploaded_file_name': last_uploaded_file_name,
            'total_data_points': 0,
            'missing_data_points': 0,
            'total_segments': 0,
            'total_abnormal_points': 0,
            'stored_start_date': stored_start_date,
            'stored_end_date': stored_end_date,
            'stored_rate_of_change': stored_rate_of_change,
            'stored_station_name': station_name
        },
        'stations': stations,
    })



def export_spikedata(request):
    # Get the start and end date from the request (or session)
    start_date_str = request.session.get('start_date', '')
    end_date_str = request.session.get('end_date', '')

    # Print the dates for testing
    print(f"Export Start Date: {start_date_str}")
    print(f"Export End Date: {end_date_str}")

    # Function to handle multiple date formats
    def parse_date(date_str):
        try:
            # Try to parse using the '%d/%m/%Y' format
            return datetime.strptime(date_str, '%d/%m/%Y')
        except ValueError:
            try:
                # If '%d/%m/%Y' fails, try '%Y-%m-%d' format
                return datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                # If both fail, return None
                return None

    # If both start and end dates are provided
    if start_date_str and end_date_str:
        # Convert start and end dates to datetime objects
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)

        if start_date and end_date:
            # Set time to 00:00 for start date and 23:59 for end date to cover full days
            start_date_str = start_date.strftime('%Y-%m-%d') + " 00:00"
            end_date_str = end_date.strftime('%Y-%m-%d') + " 23:59"

            # Filter the data between the specified date range
            spike_data_records = SpikeData.objects.filter(
                dateTime__gte=start_date_str,  # Greater than or equal to start date
                dateTime__lte=end_date_str     # Less than or equal to end date
            )
        else:
            # Invalid date format handling
            return HttpResponse('Invalid date format.')

    elif start_date_str:
        # If only start date is provided
        start_date = parse_date(start_date_str)
        if start_date:
            start_date_str = start_date.strftime('%Y-%m-%d') + " 00:00"
            spike_data_records = SpikeData.objects.filter(dateTime__gte=start_date_str)
        else:
            # Invalid date format handling
            return HttpResponse('Invalid start date format.')

    elif end_date_str:
        # If only end date is provided
        end_date = parse_date(end_date_str)
        if end_date:
            end_date_str = end_date.strftime('%Y-%m-%d') + " 23:59"
            spike_data_records = SpikeData.objects.filter(dateTime__lte=end_date_str)
        else:
            # Invalid date format handling
            return HttpResponse('Invalid end date format.')

    else:
        # If no date range is provided, return all records
        spike_data_records = SpikeData.objects.all()

    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="spike_data.csv"'

    writer = csv.writer(response)
    writer.writerow(['dateTime', 'value', 'modified_values'])  # Include modified_values in the header

    # Write each record in the filtered data
    for record in spike_data_records:
        # Ensure datetime is in %Y-%m-%d %H:%M:%S format
        try:
            # Convert the datetime to the format '%Y-%m-%d %H:%M:%S'
            record_dateTime = datetime.strptime(record.dateTime, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            # Handle any invalid date format here if necessary
            record_dateTime = None

        if record_dateTime:
            # Write the record to the CSV in the correct format
            writer.writerow([record_dateTime.strftime('%Y-%m-%d %H:%M:%S'), record.value, record.modified_value])
        else:
            # If the date format doesn't match, log an error or skip
            print(f"Skipping record due to invalid date format: {record.dateTime}")

    return response


def upload_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        
        # Load the Excel file
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        # Prepare the raw SQL query to insert the data (without specifying id)
        insert_query = """
        INSERT INTO uiapp_stationname (station_name) VALUES (%s);
        """
        
        # List to collect the data for batch insertion
        data_to_insert = []

        # Loop through the rows in the Excel file and prepare data for insertion
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            station_name = row[0]  # Assuming station names are in the first column
            if station_name:  # Check if the station name is not empty
                data_to_insert.append((station_name,))

        # Execute the raw SQL query in batches to insert the data
        with connection.cursor() as cursor:
            # Execute the insertions in batches of 2000
            batch_size = 500
            for i in range(0, len(data_to_insert), batch_size):
                cursor.executemany(insert_query, data_to_insert[i:i + batch_size])

        return HttpResponse("<h1>Data successfully uploaded!</h1>")

    return render(request, 'upload_excel.html')

def get_stations(request):
    stations = StationName.objects.values('id', 'station_name')
    return JsonResponse(list(stations), safe=False)


import openpyxl
from django.shortcuts import render
from django.http import HttpResponse
from .models import StationRecord

def upload_station_data(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']

        # Validate the uploaded file type
        if not (excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls')):
            return HttpResponse('Invalid file format. Please upload an Excel file.')

        try:
            # Load the Excel file
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Iterate through rows and save data to the database
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                station_id = row[0]
                recorded_highest_wl = row[1]
                recorded_lowest_wl = row[2]

                # Validate and save the record
                if station_id and recorded_highest_wl is not None and recorded_lowest_wl is not None:
                    StationRecord.objects.create(
                        station_id=station_id,
                        recorded_highest_wl=recorded_highest_wl,
                        recorded_lowest_wl=recorded_lowest_wl
                    )

            return HttpResponse("<h1>File uploaded and data saved successfully!</h1>")
        except Exception as e:
            return HttpResponse(f"<h1>Error processing the file: {str(e)}</h1>")

    return render(request, 'upload_station_data.html')